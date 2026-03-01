#openpyxl
import openpyxl


def readexcelIFle():
    filepth = "annual-enterprise-survey.xlsx"
    workbook= openpyxl.load_workbook(filepth)
    sheet_obj= workbook.active
    data=sheet_obj.cell(row=2, column=6).value
    print(data)
    rowno=sheet_obj.max_row
    colno=sheet_obj.max_column
    print("Total number of rows:", rowno)
    print("Total number of columns:", colno)
    #print header data adata from a single row
    for col in range(1, colno+1):
        header_data= sheet_obj.cell(row=1, column=col).value
        print(header_data, end=" ")

   #read column data
    print("\n========================Column Data========================")
    for row in range (1, 10):
        column_data= sheet_obj.cell(row=row, column=7).value
        print(column_data)
        # read complete sheet data
    print("\n========================Complete Sheet Data========================")
    for row in range(1, 10):
        for col in range(1, colno+1):
            cell_data= sheet_obj.cell(row=row, column=col).value
            print(cell_data, end=" ")
        print()

#readexcelIFle()


def writeexcelFile():
    filepth = "annual-enterprise-survey1.xlsx"
    wb= openpyxl.Workbook()
    sheet_obj= wb.active
    sheet_obj.title ="TestData"
    sheet_obj.cell(row=2, column=6).value = "New Value"
    #sheet_obj['F2'].value = "New Value"
    sheet_obj['A10'].value = "Data in A10"
    wb.save(filepth)
    print("Data written to excel file successfully")


writeexcelFile()
